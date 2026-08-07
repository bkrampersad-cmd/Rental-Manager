"""Export helpers: Excel (.xlsx), CSV, and formatted PDF reports."""
import csv
import io
from datetime import datetime
from xml.sax.saxutils import escape as xml_escape

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.utils import ImageReader
from reportlab.platypus import (
    SimpleDocTemplate,
    Table,
    TableStyle,
    Paragraph,
    Spacer,
    Image,
    HRFlowable,
    PageBreak,
    KeepTogether,
)

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_BORDER = Border(bottom=Side(style="thin", color="D1D5DB"))


# ---------------------------------------------------------------------------
# CSV
# ---------------------------------------------------------------------------

def transactions_to_csv(transactions):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Date", "Property", "Unit", "Type", "Category", "Payee", "Amount", "Account", "Notes"])
    for t in transactions:
        writer.writerow(
            [
                t["date"],
                t.get("property_name", ""),
                t.get("unit_name", "") or "",
                t["type"].title(),
                t.get("category_name", ""),
                t.get("payee", ""),
                f"{t['amount']:.2f}",
                t.get("account", ""),
                t.get("notes", "") or "",
            ]
        )
    return buf.getvalue().encode("utf-8")


def comparison_to_csv(rows):
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Property", "Income", "Expenses", "Net Cash Flow", "Cap Rate %", "Cash-on-Cash %"])
    for r in rows:
        writer.writerow(
            [
                r["name"],
                f"{r['income']:.2f}",
                f"{r['expenses']:.2f}",
                f"{r['net']:.2f}",
                f"{r['cap_rate']:.2f}" if r["cap_rate"] is not None else "",
                f"{r['coc_return']:.2f}" if r["coc_return"] is not None else "",
            ]
        )
    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# QuickBooks / Quicken migration exports — IIF (QuickBooks Desktop), QIF
# (Quicken), and a plain 3-column bank CSV (QuickBooks Online's manual bank
# transaction upload, also accepted by Quicken's CSV importer). These exist
# so someone who decides to move to one of those platforms can bring their
# transaction history with them instead of re-keying it by hand. Amount sign
# convention throughout: positive = income/deposit, negative = expense/
# payment — matching how a real bank register works.
# ---------------------------------------------------------------------------

def _iif_safe(text):
    """IIF is tab-delimited; strip anything that would break a column."""
    if not text:
        return ""
    return str(text).replace("\t", " ").replace("\n", " ").replace("\r", " ").strip()


def _qif_safe(text):
    """QIF fields are one per line; strip newlines that would break that."""
    if not text:
        return ""
    return str(text).replace("\n", " ").replace("\r", " ").strip()


def _reformat_date(iso_date):
    try:
        return datetime.strptime(iso_date, "%Y-%m-%d").strftime("%m/%d/%Y")
    except (ValueError, TypeError):
        return iso_date or ""


def transactions_to_iif(transactions):
    """QuickBooks Desktop IIF file. Creates one BANK account per distinct
    property/account-label pair, one INC or EXP account per category, and one
    CLASS per property (turn on class tracking in QuickBooks — Edit >
    Preferences > Accounting — to get a per-property P&L from these), then a
    balanced TRNS/SPL/ENDTRNS block per transaction. For a deposit, the bank
    leg is positive and the income leg is the same amount negative; for a
    check/payment it's the reverse — this is what keeps each transaction
    balanced on import."""
    bank_accounts = {}   # (property_name, account_label) -> IIF account name
    category_types = {}  # IIF category account name -> "income"/"expense"
    classes = []
    rows = []

    def bank_account_name(property_name, account_label):
        key = (property_name, account_label or "")
        if key not in bank_accounts:
            label = account_label or "Bank Account"
            name = f"{property_name} - {label}" if property_name else label
            bank_accounts[key] = _iif_safe(name)[:100] or "Bank Account"
        return bank_accounts[key]

    for t in transactions:
        prop = t.get("property_name") or ""
        if prop and prop not in classes:
            classes.append(prop)
        bank_name = bank_account_name(prop, t.get("account"))
        cat_name = _iif_safe(t.get("category_name") or "Uncategorized")[:100] or "Uncategorized"
        category_types[cat_name] = t["type"]
        rows.append((t, bank_name, cat_name, prop))

    lines = ["!ACCNT\tNAME\tACCNTTYPE"]
    for name in sorted(set(bank_accounts.values())):
        lines.append(f"ACCNT\t{name}\tBANK")
    for cat_name in sorted(category_types):
        acct_type = "INC" if category_types[cat_name] == "income" else "EXP"
        lines.append(f"ACCNT\t{cat_name}\t{acct_type}")

    if classes:
        lines.append("!CLASS\tNAME")
        for c in classes:
            lines.append(f"CLASS\t{_iif_safe(c)}")

    lines.append("!TRNS\tTRNSID\tTRNSTYPE\tDATE\tACCNT\tNAME\tCLASS\tAMOUNT\tMEMO")
    lines.append("!SPL\tSPLID\tTRNSTYPE\tDATE\tACCNT\tNAME\tCLASS\tAMOUNT\tMEMO")
    lines.append("!ENDTRNS")

    for t, bank_name, cat_name, prop in rows:
        qb_date = _reformat_date(t["date"])
        payee = _iif_safe(t.get("payee"))
        memo = _iif_safe(t.get("notes"))
        amount = t["amount"]
        is_income = t["type"] == "income"
        trns_type = "DEPOSIT" if is_income else "CHECK"
        trns_amount = amount if is_income else -amount
        spl_amount = -amount if is_income else amount
        class_field = _iif_safe(prop)
        lines.append(f"TRNS\t\t{trns_type}\t{qb_date}\t{bank_name}\t{payee}\t{class_field}\t{trns_amount:.2f}\t{memo}")
        lines.append(f"SPL\t\t{trns_type}\t{qb_date}\t{cat_name}\t{payee}\t{class_field}\t{spl_amount:.2f}\t{memo}")
        lines.append("ENDTRNS")

    return ("\r\n".join(lines) + "\r\n").encode("utf-8")


def transactions_to_qif(transactions):
    """Quicken QIF file — a single !Type:Bank register. Each transaction's
    category (and, via Quicken's Category/Class syntax, its property) is
    preserved, so nothing needs to be re-categorized by hand after
    importing."""
    lines = ["!Type:Bank"]
    for t in sorted(transactions, key=lambda x: x["date"]):
        qdate = _reformat_date(t["date"])
        amount = t["amount"] if t["type"] == "income" else -t["amount"]
        payee = t.get("payee") or (t.get("category_name") or "Transaction")
        category = t.get("category_name") or "Uncategorized"
        prop = t.get("property_name") or ""
        lines.append(f"D{qdate}")
        lines.append(f"T{amount:.2f}")
        lines.append(f"P{_qif_safe(payee)}")
        cat_field = f"{_qif_safe(category)}/{_qif_safe(prop)}" if prop else _qif_safe(category)
        lines.append(f"L{cat_field}")
        notes = t.get("notes")
        if notes:
            lines.append(f"M{_qif_safe(notes)}")
        lines.append("^")
    return ("\n".join(lines) + "\n").encode("utf-8")


def transactions_to_qb_csv(transactions):
    """Plain 3-column Date/Description/Amount CSV — the format QuickBooks
    Online's manual bank-transaction upload expects, and one Quicken's CSV
    importer also accepts. Categories aren't preserved (neither target
    column layout carries them) — use the IIF or QIF export instead if
    keeping categories intact matters more than QuickBooks Online
    compatibility specifically."""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(["Date", "Description", "Amount"])
    for t in sorted(transactions, key=lambda x: x["date"]):
        csv_date = _reformat_date(t["date"])
        bits = [b for b in [t.get("payee"), t.get("property_name")] if b]
        description = " - ".join(bits) if bits else (t.get("category_name") or "Transaction")
        amount = t["amount"] if t["type"] == "income" else -t["amount"]
        writer.writerow([csv_date, description, f"{amount:.2f}"])
    return buf.getvalue().encode("utf-8")


# ---------------------------------------------------------------------------
# Excel
# ---------------------------------------------------------------------------

def _add_letterhead_rows(ws, business, num_cols, logo_bytes=None):
    """Prepend business-name / contact rows (and logo, if provided) to a
    worksheet. Returns nothing; relies on ws.max_row so callers must add
    headers immediately after."""
    business = business or {}
    biz_name = business.get("name")
    if not biz_name and not logo_bytes:
        return

    if biz_name:
        ws.append([biz_name])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
        ws.cell(row=r, column=1).font = Font(bold=True, size=14, color="111827")

        contact_bits = [b for b in [business.get("address_line"), business.get("phone"), business.get("email")] if b]
        if contact_bits:
            ws.append([" | ".join(contact_bits)])
            r = ws.max_row
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=num_cols)
            ws.cell(row=r, column=1).font = Font(italic=True, size=10, color="6B7280")

    if logo_bytes:
        try:
            xl_img = XLImage(io.BytesIO(logo_bytes))
            max_w = 130
            if xl_img.width > max_w:
                ratio = max_w / xl_img.width
                xl_img.width = max_w
                xl_img.height = xl_img.height * ratio
            anchor_col = get_column_letter(num_cols + 2)
            ws.add_image(xl_img, f"{anchor_col}1")
        except Exception:
            pass

    ws.append([])  # spacer row


def _add_footer_row(ws, business):
    footer_text = (business or {}).get("footer")
    if not footer_text:
        return
    r = ws.max_row + 2
    ws.cell(row=r, column=1, value=footer_text).font = Font(italic=True, size=9, color="9CA3AF")


def transactions_to_excel(transactions, business=None, logo_bytes=None, title="Transactions"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Transactions"

    headers = ["Date", "Property", "Unit", "Type", "Category", "Payee", "Amount", "Account", "Notes"]
    _add_letterhead_rows(ws, business, len(headers), logo_bytes)

    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=header_row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")

    income_fill = PatternFill(start_color="E7F7EE", end_color="E7F7EE", fill_type="solid")
    expense_fill = PatternFill(start_color="FDEDEC", end_color="FDEDEC", fill_type="solid")

    total_income = 0.0
    total_expense = 0.0
    for t in transactions:
        row = [
            t["date"],
            t.get("property_name", ""),
            t.get("unit_name", "") or "",
            t["type"].title(),
            t.get("category_name", ""),
            t.get("payee", ""),
            t["amount"] if t["type"] == "income" else -t["amount"],
            t.get("account", ""),
            t.get("notes", "") or "",
        ]
        ws.append(row)
        r = ws.max_row
        fill = income_fill if t["type"] == "income" else expense_fill
        for col in range(1, len(headers) + 1):
            ws.cell(row=r, column=col).fill = fill
            ws.cell(row=r, column=col).border = THIN_BORDER
        ws.cell(row=r, column=7).number_format = '#,##0.00;[Red](#,##0.00)'
        if t["type"] == "income":
            total_income += t["amount"]
        else:
            total_expense += t["amount"]

    summary_row = ws.max_row + 2
    ws.cell(row=summary_row, column=1, value="Total Income").font = Font(bold=True)
    ws.cell(row=summary_row, column=7, value=total_income).font = Font(bold=True, color="1A9C5B")
    ws.cell(row=summary_row + 1, column=1, value="Total Expenses").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=7, value=-total_expense).font = Font(bold=True, color="D9463F")
    ws.cell(row=summary_row + 2, column=1, value="Net Cash Flow").font = Font(bold=True)
    ws.cell(row=summary_row + 2, column=7, value=total_income - total_expense).font = Font(bold=True)

    for i, header in enumerate(headers, start=1):
        max_len = max([len(str(header))] + [len(str(row[i - 1])) for row in
                     [[t["date"], t.get("property_name",""), t.get("unit_name","") or "", t["type"],
                       t.get("category_name",""), t.get("payee",""), t["amount"], t.get("account",""),
                       t.get("notes","") or ""]
                      for t in transactions]] or [10])
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 10), 40)

    ws.freeze_panes = f"A{header_row + 1}"
    _add_footer_row(ws, business)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def comparison_to_excel(rows, period_label="", business=None, logo_bytes=None):
    wb = Workbook()
    ws = wb.active
    ws.title = "Property Comparison"

    headers = ["Property", "Income", "Expenses", "Net Cash Flow", "Cap Rate %", "Cash-on-Cash %"]
    _add_letterhead_rows(ws, business, len(headers), logo_bytes)

    ws.append([f"Property Comparison — {period_label}"])
    r = ws.max_row
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers))
    ws.cell(row=r, column=1).font = Font(bold=True, size=14)

    ws.append(headers)
    header_row = ws.max_row
    for col in range(1, len(headers) + 1):
        c = ws.cell(row=header_row, column=col)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append(
            [
                r["name"],
                r["income"],
                -r["expenses"],
                r["net"],
                r["cap_rate"] if r["cap_rate"] is not None else "",
                r["coc_return"] if r["coc_return"] is not None else "",
            ]
        )
        rr = ws.max_row
        for col in (2, 3, 4):
            ws.cell(row=rr, column=col).number_format = '#,##0.00;[Red](#,##0.00)'

    for i in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 20

    _add_footer_row(ws, business)

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


# ---------------------------------------------------------------------------
# PDF
# ---------------------------------------------------------------------------

def _pdf_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportTitle", fontSize=20, leading=24, spaceAfter=4,
                               textColor=colors.HexColor("#111827"), fontName="Helvetica-Bold"))
    styles.add(ParagraphStyle(name="DocTitle", fontSize=14, leading=17, spaceAfter=4,
                               textColor=colors.HexColor("#111827"), fontName="Helvetica-Bold"))
    styles.add(ParagraphStyle(name="ReportSubtitle", fontSize=11, textColor=colors.HexColor("#6B7280"),
                               spaceAfter=16))
    styles.add(ParagraphStyle(name="SectionHeader", fontSize=13, spaceBefore=14, spaceAfter=8,
                               textColor=colors.HexColor("#111827"), fontName="Helvetica-Bold"))
    styles.add(ParagraphStyle(name="BizName", fontSize=17, leading=20,
                               textColor=colors.HexColor("#111827"), fontName="Helvetica-Bold"))
    styles.add(ParagraphStyle(name="BizContact", fontSize=9, leading=12,
                               textColor=colors.HexColor("#6B7280")))
    styles.add(ParagraphStyle(name="ReportFooter", fontSize=8.5,
                               textColor=colors.HexColor("#9CA3AF")))
    styles.add(ParagraphStyle(name="MutedNote", fontSize=9, leading=13,
                               textColor=colors.HexColor("#6B7280")))
    styles.add(ParagraphStyle(name="WarningNote", fontSize=9.5, leading=13,
                               textColor=colors.HexColor("#92400E")))
    return styles


def _build_letterhead(business, logo_bytes, styles):
    """Business-branded header block for PDF reports. Returns a list of
    flowables (empty if no business name/logo has been configured)."""
    business = business or {}
    name = business.get("name") or ""
    contact_bits = [b for b in [
        business.get("address_line"), business.get("phone"), business.get("email"), business.get("website")
    ] if b]
    contact_line = "  |  ".join(xml_escape(b) for b in contact_bits)

    if not name and not logo_bytes:
        return []

    text_flowables = []
    if name:
        text_flowables.append(Paragraph(xml_escape(name), styles["BizName"]))
    if contact_line:
        text_flowables.append(Paragraph(contact_line, styles["BizContact"]))

    elements = []
    logo_flowable = None
    if logo_bytes:
        try:
            reader = ImageReader(io.BytesIO(logo_bytes))
            iw, ih = reader.getSize()
            max_h = 0.55 * inch
            ratio = (iw / ih) if ih else 1
            logo_flowable = Image(io.BytesIO(logo_bytes), width=max_h * ratio, height=max_h)
        except Exception:
            logo_flowable = None

    if logo_flowable and text_flowables:
        t = Table([[logo_flowable, text_flowables]], colWidths=[logo_flowable.drawWidth + 14, 5.5 * inch])
        t.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LEFTPADDING", (0, 0), (0, 0), 0),
            ("LEFTPADDING", (1, 0), (1, 0), 10),
            ("TOPPADDING", (0, 0), (-1, -1), 0),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
        ]))
        elements.append(t)
    elif logo_flowable:
        elements.append(logo_flowable)
    else:
        elements.extend(text_flowables)

    elements.append(Spacer(1, 8))
    elements.append(HRFlowable(width="100%", thickness=1, color=colors.HexColor("#E5E7EB")))
    elements.append(Spacer(1, 12))
    return elements


def _build_pdf_footer(business, styles):
    footer_text = (business or {}).get("footer")
    if not footer_text:
        return []
    return [
        Spacer(1, 18),
        HRFlowable(width="100%", thickness=0.6, color=colors.HexColor("#E5E7EB")),
        Spacer(1, 6),
        Paragraph(xml_escape(footer_text), styles["ReportFooter"]),
    ]


def transactions_to_pdf(transactions, property_name="All Properties", period_label="", income_color="#1a9c5b",
                         expense_color="#d9463f", business=None, logo_bytes=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.6 * inch, bottomMargin=0.6 * inch,
                             leftMargin=0.6 * inch, rightMargin=0.6 * inch)
    styles = _pdf_styles()
    business = business or {}
    title_style = styles["DocTitle"] if business.get("name") else styles["ReportTitle"]
    story = _build_letterhead(business, logo_bytes, styles) + [
        Paragraph("Rental Property Statement", title_style),
        Paragraph(f"{property_name} &nbsp;|&nbsp; {period_label} &nbsp;|&nbsp; Generated {datetime.now().strftime('%B %d, %Y')}",
                  styles["ReportSubtitle"]),
    ]

    total_income = sum(t["amount"] for t in transactions if t["type"] == "income")
    total_expense = sum(t["amount"] for t in transactions if t["type"] == "expense")
    net = total_income - total_expense

    summary_data = [
        ["Total Income", "Total Expenses", "Net Cash Flow"],
        [f"${total_income:,.2f}", f"${total_expense:,.2f}", f"${net:,.2f}"],
    ]
    summary_table = Table(summary_data, colWidths=[2.1 * inch] * 3)
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F3F4F6")),
        ("TEXTCOLOR", (0, 1), (0, 1), colors.HexColor(income_color)),
        ("TEXTCOLOR", (1, 1), (1, 1), colors.HexColor(expense_color)),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, 1), (-1, 1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 1), (-1, 1), 14),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("INNERGRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 18))

    story.append(Paragraph("Transaction Detail", styles["SectionHeader"]))
    table_data = [["Date", "Unit", "Category", "Payee", "Amount"]]
    row_colors = []
    for t in sorted(transactions, key=lambda x: x["date"]):
        amt = t["amount"] if t["type"] == "income" else -t["amount"]
        table_data.append([
            t["date"],
            t.get("unit_name", "") or "",
            t.get("category_name", ""),
            (t.get("payee") or "")[:40],
            f"${amt:,.2f}",
        ])
        row_colors.append(income_color if t["type"] == "income" else expense_color)

    if len(table_data) == 1:
        table_data.append(["", "", "No transactions in this period", "", ""])
        row_colors.append("#6B7280")

    tbl = Table(table_data, colWidths=[0.9 * inch, 0.9 * inch, 1.6 * inch, 2.2 * inch, 1.1 * inch], repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ALIGN", (4, 0), (4, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]
    for i, c in enumerate(row_colors, start=1):
        style_cmds.append(("TEXTCOLOR", (4, i), (4, i), colors.HexColor(c)))
    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)
    story.extend(_build_pdf_footer(business, styles))

    doc.build(story)
    return buf.getvalue()


def comparison_to_pdf(rows, period_label="", income_color="#1a9c5b", expense_color="#d9463f",
                       business=None, logo_bytes=None):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.6 * inch, bottomMargin=0.6 * inch,
                             leftMargin=0.6 * inch, rightMargin=0.6 * inch)
    styles = _pdf_styles()
    business = business or {}
    title_style = styles["DocTitle"] if business.get("name") else styles["ReportTitle"]
    story = _build_letterhead(business, logo_bytes, styles) + [
        Paragraph("Property Comparison Report", title_style),
        Paragraph(f"{period_label} &nbsp;|&nbsp; Generated {datetime.now().strftime('%B %d, %Y')}",
                  styles["ReportSubtitle"]),
    ]

    table_data = [["Property", "Income", "Expenses", "Net Cash Flow", "Cap Rate", "Cash-on-Cash"]]
    for r in rows:
        table_data.append([
            r["name"],
            f"${r['income']:,.2f}",
            f"${r['expenses']:,.2f}",
            f"${r['net']:,.2f}",
            f"{r['cap_rate']:.2f}%" if r["cap_rate"] is not None else "—",
            f"{r['coc_return']:.2f}%" if r["coc_return"] is not None else "—",
        ])

    tbl = Table(table_data, colWidths=[1.5 * inch, 1.0 * inch, 1.0 * inch, 1.1 * inch, 0.9 * inch, 1.1 * inch], repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TEXTCOLOR", (1, 1), (1, -1), colors.HexColor(income_color)),
        ("TEXTCOLOR", (2, 1), (2, -1), colors.HexColor(expense_color)),
    ]
    tbl.setStyle(TableStyle(style_cmds))
    story.append(tbl)
    story.extend(_build_pdf_footer(business, styles))

    doc.build(story)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Tax Documents packet — cover page + one Schedule E-style section per
# property, ready to hand to a CPA. Numbers come from tax_report.py; this
# function only lays them out.
# ---------------------------------------------------------------------------

def _money(n):
    return f"${n:,.2f}"


def tax_packet_to_pdf(packet, business=None, logo_bytes=None, include_detail=False):
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter, topMargin=0.6 * inch, bottomMargin=0.6 * inch,
                             leftMargin=0.6 * inch, rightMargin=0.6 * inch)
    styles = _pdf_styles()
    business = business or {}
    title_style = styles["DocTitle"] if business.get("name") else styles["ReportTitle"]
    tax_year = packet["tax_year"]
    totals = packet["totals"]

    story = _build_letterhead(business, logo_bytes, styles) + [
        Paragraph(f"Tax Documents — {tax_year}", title_style),
        Paragraph(
            f"Generated {datetime.now().strftime('%B %d, %Y')} &nbsp;|&nbsp; "
            f"{len(packet['properties'])} propert{'y' if len(packet['properties']) == 1 else 'ies'} included",
            styles["ReportSubtitle"],
        ),
        Paragraph(
            "Prepared for your accountant. Figures follow IRS Schedule E (Form 1040) expense "
            "categories. This is not tax advice — please review with your CPA before filing.",
            styles["MutedNote"],
        ),
        Spacer(1, 14),
    ]

    cover_data = [
        ["Portfolio Totals", ""],
        ["Rents Received", _money(totals["rents_received"])],
        ["Total Expenses (incl. depreciation)", _money(totals["total_expenses"])],
        ["  of which Depreciation", _money(totals["depreciation"])],
        ["Net Income / (Loss)", _money(totals["net"])],
    ]
    cover_tbl = Table(cover_data, colWidths=[3.5 * inch, 2.0 * inch])
    cover_tbl.setStyle(TableStyle([
        ("SPAN", (0, 0), (1, 0)),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(cover_tbl)
    story.append(Spacer(1, 18))

    story.append(Paragraph("Properties Included", styles["SectionHeader"]))
    prop_rows = [["Property", "Address", "State", "Net Income / (Loss)"]]
    for r in packet["properties"]:
        addr = r["property_address"] or ""
        prop_rows.append([r["property_name"], addr, r["property_state"] or "", _money(r["net"])])
    prop_tbl = Table(prop_rows, colWidths=[1.7 * inch, 2.6 * inch, 0.7 * inch, 1.5 * inch], repeatRows=1)
    prop_tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F3F4F6")),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9.5),
        ("ALIGN", (3, 0), (3, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    story.append(prop_tbl)
    story.extend(_build_pdf_footer(business, styles))

    # One Schedule E-style section per property, each starting on its own page.
    for r in packet["properties"]:
        story.append(PageBreak())
        story.append(Paragraph(r["property_name"], title_style))
        addr_bits = [b for b in [r["property_address"], r["property_state"]] if b]
        story.append(Paragraph(
            (", ".join(addr_bits) if addr_bits else "") +
            f" &nbsp;|&nbsp; Tax Year {tax_year} &nbsp;|&nbsp; Schedule E (Form 1040)",
            styles["ReportSubtitle"],
        ))

        story.append(Paragraph(f"Rents Received (Line 3): {_money(r['rents_received'])}", styles["SectionHeader"]))
        story.append(Spacer(1, 6))

        line_rows = [["Line", "Expense Category", "Amount"]]
        for line in r["lines"]:
            label = line["label"] if line["key"] != "other" else "Other (list)"
            line_rows.append([str(line["number"]), label, _money(line["amount"])])
        dep = r["depreciation"]
        dep_amount_text = _money(dep["amount"])
        if dep["missing_inputs"]:
            dep_amount_text += "  (add Land Value + Placed in Service date on this property)"
        line_rows.append(["18", "Depreciation Expense or Depletion", dep_amount_text])

        line_tbl = Table(line_rows, colWidths=[0.5 * inch, 3.7 * inch, 2.3 * inch], repeatRows=1)
        line_tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9.5),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(line_tbl)

        mileage = r.get("mileage") or {}
        if mileage.get("miles"):
            story.append(Paragraph(
                f"Includes {mileage['miles']:g} mile(s) logged at the standard mileage rate "
                f"({mileage['trip_count']} trip(s), {_money(mileage['deduction'])} of the Auto and Travel "
                f"line above) — see Mileage on this property for the trip-by-trip detail.",
                styles["MutedNote"],
            ))

        story.append(Spacer(1, 10))

        summary_rows = [
            ["Total Expenses (incl. depreciation)", _money(r["total_expenses"])],
            ["Net Income / (Loss)", _money(r["net"])],
        ]
        summary_tbl = Table(summary_rows, colWidths=[4.2 * inch, 2.3 * inch])
        summary_tbl.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("LINEABOVE", (0, 0), (-1, 0), 0.6, colors.HexColor("#111827")),
            ("TOPPADDING", (0, 0), (-1, -1), 5),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ]))
        story.append(summary_tbl)
        story.append(Spacer(1, 14))

        # Informational-only figures — deliberately not part of the Schedule E
        # totals above (principal isn't deductible; improvements must be
        # capitalized/depreciated), but worth surfacing so nothing is lost.
        nd = r["not_deductible"]
        ci = r["capital_improvements"]
        adj = r.get("adjustments", {"amount": 0})
        if nd["amount"] or ci["amount"] or adj["amount"]:
            story.append(Paragraph("Tracked but not on Schedule E", styles["SectionHeader"]))
            info_rows = []
            if nd["amount"]:
                info_rows.append(["Mortgage Principal Paid", _money(nd["amount"])])
            if ci["amount"]:
                info_rows.append(["Capital Improvements", _money(ci["amount"])])
            if adj["amount"]:
                info_rows.append(["Reconciliation Adjustments", _money(adj["amount"])])
            info_tbl = Table(info_rows, colWidths=[4.2 * inch, 2.3 * inch])
            info_tbl.setStyle(TableStyle([
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("TEXTCOLOR", (0, 0), (-1, -1), colors.HexColor("#6B7280")),
                ("TOPPADDING", (0, 0), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
            ]))
            story.append(info_tbl)
            note = (
                "Principal is paid down on a loan, not an expense; capital improvements must be "
                "depreciated over time rather than deducted in the year paid. Ask your CPA about "
                "adding these to this property's depreciable basis."
            )
            if adj["amount"]:
                note += (
                    " Reconciliation adjustments correct this app's books to match a bank statement — "
                    "review each one to see if it represents real, previously-untracked income or expense "
                    "that should be re-entered under its real category before filing."
                )
            story.append(Paragraph(note, styles["MutedNote"]))
            story.append(Spacer(1, 10))

        if r["unmapped"]:
            story.append(Paragraph("&#9888; Needs Review — Not Yet Mapped to a Schedule E Line", styles["SectionHeader"]))
            story.append(Paragraph(
                "These categories haven't been assigned a Schedule E line in Settings &gt; Categories, "
                "so they are NOT included in the totals above. Map them and regenerate this report before filing.",
                styles["WarningNote"],
            ))
            warn_rows = [["Category", "Amount"]] + [
                [u["category_name"], _money(u["amount"])] for u in r["unmapped"]
            ]
            warn_tbl = Table(warn_rows, colWidths=[4.2 * inch, 2.3 * inch], repeatRows=1)
            warn_tbl.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FEF3C7")),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (1, 0), (1, -1), "RIGHT"),
                ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#FDE68A")),
                ("TOPPADDING", (0, 0), (-1, -1), 5),
                ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
            ]))
            story.append(warn_tbl)
            story.append(Spacer(1, 10))

        if include_detail:
            detail_rows = [["Line", "Date", "Payee", "Amount"]]
            detail_source = [("3 — Rents Received", r.get("rents_transactions", []))]
            for line in r["lines"]:
                if line.get("transactions"):
                    detail_source.append((f"{line['number']} — {line['label']}", line["transactions"]))
            for label, txns in detail_source:
                for tx in txns:
                    detail_rows.append([label, tx["date"], (tx.get("payee") or "")[:40], _money(tx["amount"])])
            if len(detail_rows) > 1:
                story.append(Paragraph("Transaction Detail", styles["SectionHeader"]))
                detail_tbl = Table(detail_rows, colWidths=[2.0 * inch, 0.9 * inch, 2.4 * inch, 1.2 * inch], repeatRows=1)
                detail_tbl.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F2937")),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                    ("ALIGN", (3, 0), (3, -1), "RIGHT"),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#E5E7EB")),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F9FAFB")]),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]))
                story.append(detail_tbl)

        story.extend(_build_pdf_footer(business, styles))

    doc.build(story)
    return buf.getvalue()
