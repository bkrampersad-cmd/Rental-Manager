"""Generic transaction importer — CSV, Excel (.xlsx), and PDF statements.

Works with CSV/Excel exports from most banks and credit unions (Chase, BofA,
Wells Fargo, etc. all export a variant of: date, description, amount[,
debit, credit]) as well as richer ledger-style exports from other property/
accounting software that may include property, category, type, account, and
notes columns. The user maps each of the app's own fields to a column (or
leaves it unmapped), and unmapped columns/rows are simply left out of the
import.

PDF statements are supported too: text-based (digital) PDFs are parsed
directly; pages that turn out to be scans/photos (little to no extractable
text) fall back to OCR via Tesseract, if it's installed on this machine.
OCR is inherently less reliable than reading real text, so callers should
surface the `warnings` list this module returns and tell the user to
double-check anything that came from an OCR'd page before importing it.
"""
import csv
import hashlib
import io
import os
import re
import sys

from dateutil import parser as date_parser

# ---------------------------------------------------------------------------
# Optional bundled Tesseract OCR — if a tesseract-bin\ folder was placed next
# to the app (see BUILD INSTRUCTIONS.txt: install Tesseract once, copy its
# Program Files\Tesseract-OCR folder in as tesseract-bin\ before running
# build_exe.bat), point pytesseract straight at those binaries instead of
# relying on a separate system-wide install. Entirely optional — if that
# folder isn't there, this is a no-op and OCR behaves exactly as it always
# has (pytesseract falls back to whatever "tesseract" resolves to on PATH).
# ---------------------------------------------------------------------------
_IS_FROZEN = getattr(sys, "frozen", False)
if _IS_FROZEN:
    _RESOURCE_DIR = getattr(sys, "_MEIPASS", os.path.dirname(sys.executable))
else:
    _RESOURCE_DIR = os.path.dirname(os.path.abspath(__file__))

_BUNDLED_TESSERACT_DIR = os.path.join(_RESOURCE_DIR, "tesseract-bin")
_BUNDLED_TESSERACT_EXE = os.path.join(_BUNDLED_TESSERACT_DIR, "tesseract.exe")
_tesseract_bundle_checked = False


def _configure_bundled_tesseract():
    """One-time check, safe to call repeatedly. Wires pytesseract up to the
    bundled binary + tessdata if present, then actually confirms it runs
    (not just that the file exists) before committing to it — if the
    bundle turns out to be broken (missing a DLL, wrong architecture,
    incomplete copy), this falls back to plain "tesseract" so a separate
    system-wide install still gets a chance, exactly like if there'd been
    no bundle at all. If neither works, OCR just degrades gracefully
    later (see _ocr_page_text) same as it always has."""
    global _tesseract_bundle_checked
    if _tesseract_bundle_checked:
        return
    _tesseract_bundle_checked = True
    if not os.path.isfile(_BUNDLED_TESSERACT_EXE):
        return
    try:
        import pytesseract
    except ImportError:
        return
    pytesseract.pytesseract.tesseract_cmd = _BUNDLED_TESSERACT_EXE
    tessdata_dir = os.path.join(_BUNDLED_TESSERACT_DIR, "tessdata")
    if os.path.isdir(tessdata_dir):
        os.environ["TESSDATA_PREFIX"] = tessdata_dir
    try:
        pytesseract.get_tesseract_version()  # confirms the bundle actually runs
    except Exception:
        pytesseract.pytesseract.tesseract_cmd = "tesseract"
        os.environ.pop("TESSDATA_PREFIX", None)


def get_ocr_status():
    """Small diagnostic used by Settings so you can confirm OCR is actually
    working (and whether it's using the tesseract-bin\\ bundle or a
    separate system-wide install) without needing a real scanned PDF on
    hand to test with. Actually invokes the Tesseract binary (not just a
    file-existence check), so a bundle that's missing a DLL or otherwise
    broken shows up here rather than only failing later on a real import."""
    try:
        import pytesseract
    except ImportError:
        return {"available": False, "source": None, "version": None,
                 "error": "The pytesseract Python package isn't installed."}

    _configure_bundled_tesseract()
    source = "bundled" if pytesseract.pytesseract.tesseract_cmd == _BUNDLED_TESSERACT_EXE else "system"

    try:
        version = str(pytesseract.get_tesseract_version())
    except Exception as exc:
        return {"available": False, "source": source, "version": None, "error": str(exc)}

    return {"available": True, "source": source, "version": version, "error": None}

CANDIDATE_DATE_HEADERS = ["date", "transaction date", "posting date", "posted date"]
CANDIDATE_DESC_HEADERS = ["description", "memo", "payee", "name", "details"]
CANDIDATE_AMOUNT_HEADERS = ["amount", "transaction amount"]
CANDIDATE_DEBIT_HEADERS = ["debit", "withdrawal", "withdrawals", "money out"]
CANDIDATE_CREDIT_HEADERS = ["credit", "deposit", "deposits", "money in"]
CANDIDATE_PROPERTY_HEADERS = ["property", "property name", "unit", "building", "address"]
CANDIDATE_CATEGORY_HEADERS = ["category", "expense category", "account category", "class"]
CANDIDATE_TYPE_HEADERS = ["type", "transaction type", "income/expense", "debit/credit"]
CANDIDATE_ACCOUNT_HEADERS = ["account", "bank account", "source account"]
CANDIDATE_NOTES_HEADERS = ["notes", "comment", "comments"]

# Values recognized in an explicit "type" column, mapped to our internal
# income/expense vocabulary. Anything not listed here falls back to
# inferring type from the sign of the amount.
TYPE_VALUE_MAP = {
    "income": "income", "credit": "income", "deposit": "income", "in": "income",
    "expense": "expense", "debit": "expense", "withdrawal": "expense", "out": "expense",
}


# ---------------------------------------------------------------------------
# Bank sub-account detection — some combined exports/statements bundle
# multiple accounts together with a section header announcing each one
# (e.g. "Checking Account ...1234" followed by a block of transactions,
# then "Savings Account ...5678" followed by more). We look for those
# headers and use them to tag the transactions under them with an account
# label, so a single import can still tell accounts apart.
# ---------------------------------------------------------------------------

ACCOUNT_SECTION_RE = re.compile(
    r"\b(checking|savings|money\s*market|business\s*checking|share\s*draft|certificate of deposit)\b",
    re.I,
)
ACCOUNT_TAIL_DIGITS_RE = re.compile(
    r"(?:x{2,}|\*{2,}|\.{2,}|#\s*|ending in\s*|acct\.?\s*#?\s*)(\d{3,8})", re.I
)


def _account_label_from_text(text):
    """If `text` looks like a bank-statement section header naming an
    account (e.g. "Checking Account ending in 1234"), return a normalized
    label like "Checking ...1234" (or just "Checking" with no digits found)
    — otherwise None."""
    text = (text or "").strip()
    if not text or len(text) > 80:
        return None
    m = ACCOUNT_SECTION_RE.search(text)
    if not m:
        return None
    kind = re.sub(r"\s+", " ", m.group(1)).title()
    tail = ACCOUNT_TAIL_DIGITS_RE.search(text)
    if tail:
        return f"{kind} ...{tail.group(1)[-4:]}"
    return kind


def _apply_account_sections_to_rows(rows):
    """Given rows (list of list-of-str) that may contain "section header"
    rows announcing a new bank account — a row with exactly one populated
    cell whose text names an account type — strip those rows out and
    return (remaining_rows, account_labels), where account_labels[i] is
    the detected account for remaining_rows[i] (or None if none seen yet).
    Returns (rows, None) unchanged if no section headers were found at all,
    so files that don't use this pattern are completely unaffected."""
    current = None
    found_any = False
    kept_rows = []
    labels = []
    for r in rows:
        non_empty = [c for c in r if c and c.strip()]
        if len(non_empty) == 1:
            label = _account_label_from_text(non_empty[0])
            if label:
                current = label
                found_any = True
                continue  # this row is a header, not a transaction
        kept_rows.append(r)
        labels.append(current)
    if not found_any:
        return rows, None
    return kept_rows, labels


def _guess(headers_lower, candidates):
    for c in candidates:
        for i, h in enumerate(headers_lower):
            if h == c:
                return i
    for c in candidates:
        for i, h in enumerate(headers_lower):
            if c in h:
                return i
    return None


def _build_guess(headers, data_rows):
    """Shared column-guessing logic for CSV/Excel/PDF alike: match header
    names first, then fall back to inspecting the actual cell content —
    the latter matters most for PDFs, where extracted "headers" are often
    just placeholders like "Column 1" rather than real field names."""
    headers_lower = [h.strip().lower() for h in headers]
    guess = {
        "date_col": _guess(headers_lower, CANDIDATE_DATE_HEADERS),
        "description_col": _guess(headers_lower, CANDIDATE_DESC_HEADERS),
        "amount_col": _guess(headers_lower, CANDIDATE_AMOUNT_HEADERS),
        "debit_col": _guess(headers_lower, CANDIDATE_DEBIT_HEADERS),
        "credit_col": _guess(headers_lower, CANDIDATE_CREDIT_HEADERS),
        "property_col": _guess(headers_lower, CANDIDATE_PROPERTY_HEADERS),
        "category_col": _guess(headers_lower, CANDIDATE_CATEGORY_HEADERS),
        "type_col": _guess(headers_lower, CANDIDATE_TYPE_HEADERS),
        "account_col": _guess(headers_lower, CANDIDATE_ACCOUNT_HEADERS),
        "notes_col": _guess(headers_lower, CANDIDATE_NOTES_HEADERS),
    }

    def taken():
        return {v for v in guess.values() if v is not None}

    # If we couldn't find a single amount column, rely on debit/credit split instead.
    if guess["amount_col"] is None and guess["debit_col"] is None and guess["credit_col"] is None:
        for i, h in enumerate(headers_lower):
            if i in taken():
                continue
            sample_vals = [r[i] for r in data_rows[:10] if len(r) > i]
            if sample_vals and all(_looks_numeric(v) for v in sample_vals if v.strip()):
                guess["amount_col"] = i
                break

    # Content-based fallbacks — mainly for PDF-extracted tables, where
    # headers are often synthetic placeholders rather than real field names.
    if guess["date_col"] is None:
        best_i, best_score = None, 0
        for i in range(len(headers)):
            if i in taken():
                continue
            sample_vals = [r[i] for r in data_rows[:15] if len(r) > i and r[i].strip()]
            if not sample_vals:
                continue
            hits = sum(1 for v in sample_vals if _looks_like_date(v))
            score = hits / len(sample_vals)
            if score > 0.6 and score > best_score:
                best_i, best_score = i, score
        guess["date_col"] = best_i

    if guess["amount_col"] is None and guess["debit_col"] is None and guess["credit_col"] is None:
        for i, h in enumerate(headers_lower):
            if i in taken():
                continue
            sample_vals = [r[i] for r in data_rows[:15] if len(r) > i and r[i].strip()]
            if sample_vals and all(_looks_numeric(v) for v in sample_vals):
                guess["amount_col"] = i
                break

    if guess["description_col"] is None:
        best_i, best_len = None, 0
        for i in range(len(headers)):
            if i in taken():
                continue
            sample_vals = [r[i] for r in data_rows[:15] if len(r) > i and r[i].strip()]
            if not sample_vals:
                continue
            if all(_looks_numeric(v) or _looks_like_date(v) for v in sample_vals):
                continue  # not a description column
            avg_len = sum(len(v) for v in sample_vals) / len(sample_vals)
            if avg_len > best_len:
                best_i, best_len = i, avg_len
        guess["description_col"] = best_i

    return guess


def sniff_csv(file_bytes):
    """Parse the raw CSV bytes and return headers, sample rows, and a best-guess mapping."""
    text = file_bytes.decode("utf-8-sig", errors="replace")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
    except csv.Error:
        dialect = csv.excel

    reader = csv.reader(io.StringIO(text), dialect)
    rows = [r for r in reader if any(cell.strip() for cell in r)]
    if not rows:
        return {"headers": [], "rows": [], "guess": {}, "row_count": 0, "warnings": []}

    rows, section_labels = _apply_account_sections_to_rows(rows)
    headers = rows[0]
    data_rows = rows[1:]
    warnings = []
    if section_labels is not None:
        headers = list(headers) + ["Account (detected)"]
        row_labels = section_labels[1:]
        data_rows = [list(r) + [row_labels[i] or ""] for i, r in enumerate(data_rows)]
        warnings.append(
            "This file appears to bundle more than one bank account together — an "
            "\"Account (detected)\" column was added showing which account each row "
            "came from, based on section headers found in the file."
        )
    guess = _build_guess(headers, data_rows)

    return {
        "headers": headers,
        "rows": data_rows[:5000],  # cap for safety
        "guess": guess,
        "row_count": len(data_rows),
        "warnings": warnings,
    }


def sniff_xlsx(file_bytes):
    """Parse an uploaded .xlsx workbook the same way sniff_csv parses a CSV
    — first row is treated as headers, everything after is data. Uses the
    sheet with the most non-empty rows if the workbook has more than one."""
    from openpyxl import load_workbook

    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    best_sheet, best_count = None, -1
    for ws in wb.worksheets:
        count = ws.max_row or 0
        if count > best_count:
            best_sheet, best_count = ws, count

    if best_sheet is None:
        return {"headers": [], "rows": [], "guess": {}, "row_count": 0, "warnings": []}

    raw_rows = []
    for row in best_sheet.iter_rows(values_only=True):
        cells = ["" if v is None else _cell_to_str(v) for v in row]
        if any(c.strip() for c in cells):
            raw_rows.append(cells)

    if not raw_rows:
        return {"headers": [], "rows": [], "guess": {}, "row_count": 0, "warnings": []}

    raw_rows, section_labels = _apply_account_sections_to_rows(raw_rows)
    headers = raw_rows[0]
    data_rows = raw_rows[1:]
    warnings = []
    if section_labels is not None:
        headers = list(headers) + ["Account (detected)"]
        row_labels = section_labels[1:]
        data_rows = [list(r) + [row_labels[i] or ""] for i, r in enumerate(data_rows)]
        warnings.append(
            "This file appears to bundle more than one bank account together — an "
            "\"Account (detected)\" column was added showing which account each row "
            "came from, based on section headers found in the file."
        )
    guess = _build_guess(headers, data_rows)

    return {
        "headers": headers,
        "rows": data_rows[:5000],
        "guess": guess,
        "row_count": len(data_rows),
        "warnings": warnings,
    }


def _cell_to_str(v):
    import datetime as _dt
    if isinstance(v, (_dt.date, _dt.datetime)):
        return v.date().isoformat() if isinstance(v, _dt.datetime) else v.isoformat()
    if isinstance(v, float):
        # Preserve whole numbers cleanly (e.g. 1500.0 -> "1500") while still
        # keeping real decimals (e.g. 1500.5 -> "1500.5").
        return str(int(v)) if v.is_integer() else str(v)
    return str(v).strip()


# ---------------------------------------------------------------------------
# PDF statements — text extraction with an OCR fallback for scanned pages
# ---------------------------------------------------------------------------

_DATE_LINE_RE = re.compile(
    r"^\s*(\d{1,2}[/\-]\d{1,2}(?:[/\-]\d{2,4})?|"
    r"[A-Za-z]{3,9}\.?\s+\d{1,2},?(?:\s+\d{2,4})?)\s+(.+)$"
)
_AMOUNT_TOKEN_RE = re.compile(r"\(?-?\$?\d[\d,]*\.\d{2}\)?-?")


def _looks_like_date(v):
    v = (v or "").strip()
    if not v or len(v) > 40:
        return False
    try:
        date_parser.parse(v, fuzzy=False)
        return True
    except (ValueError, OverflowError, TypeError):
        return False


def _parse_statement_line(line):
    """Best-effort split of one line of extracted statement text into
    (date_str, description, amount_str), for statements that are just text
    rather than a clean grid table. Returns None if the line doesn't look
    like a transaction row at all."""
    line = line.strip()
    if not line:
        return None
    m = _DATE_LINE_RE.match(line)
    if not m:
        return None
    date_str, rest = m.group(1), m.group(2)

    amounts = _AMOUNT_TOKEN_RE.findall(rest)
    if not amounts:
        return None
    amount_str = amounts[-1]  # last number on the line — balance columns come after the amount
    if len(amounts) >= 2:
        amount_str = amounts[-2]  # if a running balance is present, it's last; amount is second-to-last

    desc = rest[:rest.rfind(amount_str)].strip(" -\t")
    if not desc:
        desc = rest.strip()
    return date_str, desc, amount_str


def _extract_pdf_page_text(page):
    """Returns extracted text for a pdfplumber page, trying its normal text
    layer first."""
    try:
        return page.extract_text() or ""
    except Exception:
        return ""


def _ocr_page_text(pdfium_page, dpi=250):
    """Rasterizes a pypdfium2 page and runs Tesseract OCR on it. Returns ""
    (rather than raising) if Tesseract isn't installed on this machine, so
    callers can degrade gracefully with a clear message instead of a crash."""
    try:
        import pytesseract
    except ImportError:
        return None  # OCR libraries not installed at all

    _configure_bundled_tesseract()

    try:
        scale = dpi / 72  # pypdfium2 renders at 72 dpi by default; scale up for OCR quality
        bitmap = pdfium_page.render(scale=scale)
        img = bitmap.to_pil()
        return pytesseract.image_to_string(img)
    except Exception:
        return None  # Tesseract engine itself missing/misconfigured, or OCR failed


def sniff_pdf(file_bytes):
    """Extract transaction-looking rows from a PDF bank/credit-card
    statement. Tries pdfplumber's table detection first (works well for
    statements with real gridlines), then falls back to a line-by-line
    date/amount heuristic for statements that are just laid-out text.
    Pages with little to no extractable text are treated as scans/photos
    and OCR'd via Tesseract if it's available; a warning is always
    returned when OCR was used, since it can misread characters."""
    import pypdfium2 as pdfium  # permissively-licensed (Apache-2.0/BSD) PDF rasterizer, used for OCR fallback
    import pdfplumber

    warnings = []
    ocr_pages = 0
    unreadable_pages = 0
    table_rows = []
    line_rows = []
    current_account = None  # tracked across the whole document, not just one page
    found_account_sections = False

    doc = pdfium.PdfDocument(file_bytes)
    with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
        for page_index, page in enumerate(pdf.pages):
            text = _extract_pdf_page_text(page)
            is_scanned = len(text.strip()) < 20

            if is_scanned:
                ocr_text = _ocr_page_text(doc[page_index])
                if ocr_text is None:
                    unreadable_pages += 1
                    continue
                ocr_pages += 1
                text = ocr_text
            else:
                # Still try structured table extraction on genuine text pages.
                try:
                    tables = page.extract_tables()
                except Exception:
                    tables = []
                for t in tables:
                    for row in t:
                        cells = [(c or "").strip() for c in row]
                        if any(cells):
                            table_rows.append(cells)

            for raw_line in text.splitlines():
                parsed = _parse_statement_line(raw_line)
                if parsed:
                    line_rows.append(list(parsed) + [current_account or ""])
                    continue
                # Not a transaction line — see if it's an account section
                # header instead (e.g. "Checking Account ending in 1234").
                label = _account_label_from_text(raw_line)
                if label:
                    current_account = label
                    found_account_sections = True

    doc.close()

    if table_rows:
        table_rows, table_section_labels = _apply_account_sections_to_rows(table_rows)
        if table_section_labels is not None:
            found_account_sections = True
            table_rows = [list(r) + [table_section_labels[i] or ""] for i, r in enumerate(table_rows)]

    if not found_account_sections:
        # Strip the account slot we provisionally attached to every line row
        # — no section headers were actually found, so keep the original
        # 3-column (Date/Description/Amount) shape unchanged.
        line_rows = [r[:3] for r in line_rows]

    if ocr_pages:
        warnings.append(
            f"{ocr_pages} page(s) of this PDF appeared to be a scan/photo rather than "
            "selectable text, so they were read using OCR. OCR can misread digits and "
            "characters — please carefully check the dates, amounts, and descriptions "
            "below against the original statement before importing."
        )
    if unreadable_pages:
        warnings.append(
            f"{unreadable_pages} page(s) looked like a scan/photo but couldn't be read: "
            "OCR (Tesseract) isn't installed on this machine. Install Tesseract OCR to "
            "read scanned statements, or use a digital/text-based PDF instead."
        )

    # Prefer real detected tables (more columns = more useful mapping
    # options) over the 3-column line heuristic, when both picked something up.
    if table_rows and len(table_rows) >= len(line_rows):
        # Treat the first row as headers only if it doesn't look like data
        # itself (i.e. doesn't parse as a date or a plain number).
        first = table_rows[0]
        looks_like_header = not any(_looks_like_date(c) or _looks_numeric(c) for c in first if c)
        if looks_like_header:
            headers, data_rows = first, table_rows[1:]
        else:
            headers = [f"Column {i + 1}" for i in range(len(first))]
            data_rows = table_rows
        if found_account_sections:
            headers = list(headers[:-1]) + ["Account (detected)"]
    elif line_rows:
        headers = ["Date", "Description", "Amount", "Account (detected)"] if found_account_sections \
            else ["Date", "Description", "Amount"]
        data_rows = line_rows
    else:
        headers, data_rows = [], []
        warnings.append(
            "Couldn't find anything that looked like transaction rows in this PDF. "
            "It may use a layout this importer doesn't recognize — a CSV or Excel "
            "export from your bank will generally work more reliably than a PDF."
        )

    if found_account_sections and headers:
        warnings.append(
            "This PDF appears to bundle more than one bank account together — an "
            "\"Account (detected)\" column was added showing which account each row "
            "came from, based on section headers found in the statement."
        )

    guess = _build_guess(headers, data_rows) if headers else {}

    return {
        "headers": headers,
        "rows": data_rows[:5000],
        "guess": guess,
        "row_count": len(data_rows),
        "warnings": warnings,
    }


def sniff_file(filename, file_bytes):
    """Dispatches to the right sniffer based on file extension."""
    ext = (filename.rsplit(".", 1)[-1].lower() if filename and "." in filename else "")
    if ext in ("xlsx", "xlsm"):
        return sniff_xlsx(file_bytes)
    if ext == "pdf":
        return sniff_pdf(file_bytes)
    if ext == "iif":
        return sniff_iif(file_bytes)
    if ext == "qif":
        return sniff_qif(file_bytes)
    return sniff_csv(file_bytes)


# ---------------------------------------------------------------------------
# QuickBooks Desktop (.iif) and Quicken (.qif) — for migrating IN from one
# of those platforms, the reverse of the exports in exports.py. Both parsers
# below produce the exact same {headers, rows, guess, row_count, warnings}
# shape sniff_csv/sniff_xlsx/sniff_pdf return, with every column already
# correctly guessed (there's no ambiguity to resolve — these are structured
# formats, not freeform bank exports) — so the existing column-mapping
# preview screen, build_transactions(), and _commit_parsed_rows() in app.py
# all work completely unchanged; the user just gets a mapping screen that's
# already filled in correctly, and can still adjust it before importing.
# ---------------------------------------------------------------------------

_IIF_ROW_HEADERS = ["Date", "Payee", "Category", "Property (Class)", "Amount", "Type", "Memo", "Bank Account"]
_QIF_ROW_HEADERS = ["Date", "Payee", "Category", "Property (Class)", "Amount", "Type", "Memo"]


def _structured_guess(headers):
    """Column-guess mapping for the fixed, always-in-this-order columns
    _IIF_ROW_HEADERS/_QIF_ROW_HEADERS produce above."""
    idx = {h: i for i, h in enumerate(headers)}
    return {
        "date_col": idx.get("Date"),
        "description_col": idx.get("Payee"),
        "amount_col": idx.get("Amount"),
        "debit_col": None,
        "credit_col": None,
        "property_col": idx.get("Property (Class)"),
        "category_col": idx.get("Category"),
        "type_col": idx.get("Type"),
        "account_col": idx.get("Bank Account"),
        "notes_col": idx.get("Memo"),
    }


def sniff_iif(file_bytes):
    """Reads a QuickBooks Desktop IIF export's bank-register transactions
    (TRNS/SPL/ENDTRNS blocks) — the kind produced by exporting a specific
    bank account's register, which is what exports.py's own IIF export
    also produces, and what QuickBooks Desktop itself writes when you
    export a register or transaction list. Reads any TRNSTYPE (DEPOSIT,
    CHECK, PAYMENT, GENERAL JOURNAL, etc.) the same way, using plain
    double-entry sign logic rather than special-casing each type: whichever
    line is the TRNS line is treated as the bank/cash side (its ACCNT
    becomes the row's "Bank Account", its sign decides income vs. expense),
    and the first SPL line is treated as the category side. A transaction
    split across more than one SPL line only keeps the first split's
    category — the rest of that split's amount isn't lost (it's still part
    of the total), but its own separate categorization is, so those rows
    are flagged in a warning for a manual double-check after importing.
    Invoice/bill/estimate detail elsewhere in a full company-file IIF
    export (anything outside a TRNS/SPL/ENDTRNS block) is ignored — this
    only reads actual cash-basis bank transactions, matching what this app
    tracks."""
    text = file_bytes.decode("utf-8-sig", errors="replace")
    lines = text.splitlines()
    n = len(lines)

    trns_cols = None
    spl_cols = None
    blocks = []
    multi_split_count = 0

    i = 0
    while i < n:
        parts = lines[i].split("\t")
        tag = parts[0].strip() if parts else ""
        if tag == "!TRNS":
            trns_cols = parts
            i += 1
        elif tag == "!SPL":
            spl_cols = parts
            i += 1
        elif tag == "TRNS":
            trns_fields = dict(zip(trns_cols, parts)) if trns_cols else {}
            spl_fields_list = []
            j = i + 1
            while j < n:
                jparts = lines[j].split("\t")
                jtag = jparts[0].strip() if jparts else ""
                if jtag == "SPL":
                    spl_fields_list.append(dict(zip(spl_cols, jparts)) if spl_cols else {})
                    j += 1
                elif jtag == "ENDTRNS":
                    j += 1
                    break
                else:
                    break  # malformed block — stop scanning rather than misreading unrelated lines
            i = j
            if len(spl_fields_list) > 1:
                multi_split_count += 1
            blocks.append((trns_fields, spl_fields_list))
        else:
            i += 1

    rows = []
    for trns_fields, spl_fields_list in blocks:
        try:
            amount = float((trns_fields.get("AMOUNT") or "0").replace(",", ""))
        except ValueError:
            amount = 0.0
        first_spl = spl_fields_list[0] if spl_fields_list else {}
        category = first_spl.get("ACCNT", "")
        prop = trns_fields.get("CLASS", "") or first_spl.get("CLASS", "")
        rows.append([
            trns_fields.get("DATE", ""),
            trns_fields.get("NAME", ""),
            category,
            prop,
            f"{abs(amount):.2f}",
            "income" if amount >= 0 else "expense",
            trns_fields.get("MEMO", ""),
            trns_fields.get("ACCNT", ""),
        ])

    warnings = []
    if not rows:
        warnings.append(
            "No bank transactions (TRNS/SPL blocks) were found in this IIF file. Make sure it's a "
            "register or transaction-list export for one bank account, not a full company-file backup."
        )
    if multi_split_count:
        warnings.append(
            f"{multi_split_count} transaction(s) in this file were split across more than one category "
            "— only the first category of each was kept here; the total amount is still correct, but "
            "double-check those rows' categories after importing."
        )

    return {
        "headers": _IIF_ROW_HEADERS,
        "rows": rows[:5000],
        "guess": _structured_guess(_IIF_ROW_HEADERS),
        "row_count": len(rows),
        "warnings": warnings,
    }


def sniff_qif(file_bytes):
    """Reads a Quicken QIF register export (!Type:Bank / !Type:Cash /
    !Type:CCard, then D/T/P/L/M/^ records per transaction). Understands the
    Category/Class syntax (an L line like "Rent Income/123 Main St") that
    this app's own QIF export uses to carry a transaction's property along
    — but reads a plain Quicken-generated QIF (no "/Class" part) exactly as
    well, it just leaves the property column blank for those rows so the
    mapping screen's normal "guess property from payee" fallback can take
    over instead."""
    text = file_bytes.decode("utf-8-sig", errors="replace")
    lines = text.splitlines()

    records = []
    cur = {}
    for raw in lines:
        line = raw.rstrip("\r\n")
        if not line or line.startswith("!"):
            continue  # blank line, or a !Type:... header — not transaction data
        code, rest = line[0], line[1:]
        if code == "^":
            if cur:
                records.append(cur)
            cur = {}
        elif code in ("T", "U"):
            cur["amount"] = rest.strip()
        elif code == "D":
            cur["date"] = rest.strip()
        elif code == "P":
            cur["payee"] = rest.strip()
        elif code == "L":
            if "/" in rest:
                cat, cls = rest.split("/", 1)
            else:
                cat, cls = rest, ""
            cur["category"] = cat.strip()
            cur["property"] = cls.strip()
        elif code == "M":
            cur["memo"] = rest.strip()
        elif code == "N" and "memo" not in cur:
            cur["memo"] = rest.strip()  # check/reference number — kept only if there's no real memo
        # other codes (C cleared status, A address lines, S/E splits, etc.) intentionally ignored
    if cur:  # tolerate a file missing its final "^" terminator
        records.append(cur)

    rows = []
    for r in records:
        try:
            amount = float((r.get("amount") or "0").replace(",", ""))
        except ValueError:
            amount = 0.0
        rows.append([
            r.get("date", ""),
            r.get("payee", ""),
            r.get("category", ""),
            r.get("property", ""),
            f"{abs(amount):.2f}",
            "income" if amount >= 0 else "expense",
            r.get("memo", ""),
        ])

    warnings = []
    if not rows:
        warnings.append("No transactions were found in this QIF file.")

    return {
        "headers": _QIF_ROW_HEADERS,
        "rows": rows[:5000],
        "guess": _structured_guess(_QIF_ROW_HEADERS),
        "row_count": len(rows),
        "warnings": warnings,
    }


def _looks_numeric(v):
    v = v.replace("$", "").replace(",", "").replace("(", "-").replace(")", "").strip()
    try:
        float(v)
        return True
    except ValueError:
        return False


def _parse_amount(v):
    if v is None:
        return 0.0
    v = str(v).replace("$", "").replace(",", "").strip()
    negative = v.startswith("(") and v.endswith(")")
    v = v.strip("()")
    if not v:
        return 0.0
    try:
        val = float(v)
    except ValueError:
        return 0.0
    return -val if negative else val


def build_transactions(rows, mapping):
    """Convert mapped raw rows into normalized transaction dicts.

    `rows` should already exclude any rows the user chose to skip — this
    function has no notion of row exclusion, only column mapping.

    mapping keys (all optional except date_col/description_col, indexes may
    be None): date_col, description_col, amount_col, debit_col, credit_col,
    property_col, category_col, type_col, account_col, notes_col.

    Returns a list of dicts with date, payee, amount (positive), type
    (income/expense), import_hash, and — when mapped — raw property_name,
    category_name, account, notes strings for the caller to resolve against
    the database.
    """
    date_col = mapping.get("date_col")
    desc_col = mapping.get("description_col")
    amount_col = mapping.get("amount_col")
    debit_col = mapping.get("debit_col")
    credit_col = mapping.get("credit_col")
    property_col = mapping.get("property_col")
    category_col = mapping.get("category_col")
    type_col = mapping.get("type_col")
    account_col = mapping.get("account_col")
    notes_col = mapping.get("notes_col")

    out = []
    for r in rows:
        def cell(i):
            return r[i].strip() if i is not None and i < len(r) else ""

        raw_date = cell(date_col)
        payee = cell(desc_col)
        if not raw_date and not payee:
            continue

        try:
            parsed_date = date_parser.parse(raw_date, fuzzy=True).date().isoformat()
        except (ValueError, OverflowError, OSError):
            continue

        signed_amount = None
        if amount_col is not None:
            signed_amount = _parse_amount(cell(amount_col))
        else:
            debit = _parse_amount(cell(debit_col)) if debit_col is not None else 0.0
            credit = _parse_amount(cell(credit_col)) if credit_col is not None else 0.0
            signed_amount = credit - abs(debit)

        if signed_amount is None:
            continue

        txn_type = None
        if type_col is not None:
            raw_type = cell(type_col).strip().lower()
            txn_type = TYPE_VALUE_MAP.get(raw_type)
        if txn_type is None:
            txn_type = "income" if signed_amount >= 0 else "expense"

        amount = abs(signed_amount)
        h = hashlib.sha256(f"{parsed_date}|{payee}|{amount}|{txn_type}".encode()).hexdigest()

        row_out = {
            "date": parsed_date,
            "payee": payee,
            "amount": round(amount, 2),
            "type": txn_type,
            "import_hash": h,
        }
        if property_col is not None:
            row_out["property_name"] = cell(property_col)
        if category_col is not None:
            row_out["category_name"] = cell(category_col)
        if account_col is not None:
            row_out["account"] = cell(account_col)
        if notes_col is not None:
            row_out["notes"] = cell(notes_col)

        out.append(row_out)
    return out
